from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="534AB7")
blue_font = Font(color="0000FF")
black_font = Font(color="000000")
note_font = Font(color="595959", italic=True, size=10)
title_font = Font(bold=True, size=14, color="534AB7")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def hdr(ws, row, n):
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

def data_border(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = thin_border
            ws.cell(row=r, column=c).font = black_font

# ---------------- Sheet 1: 模型定价假设 ----------------
ws1 = wb.active
ws1.title = "模型定价假设"
ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 20
ws1.column_dimensions["C"].width = 50
ws1["A1"] = "Qwen3.7-Plus 定价与成本假设"
ws1["A1"].font = title_font
ws1.merge_cells("A1:C1")

rows = [
    ("模型名称", "Qwen3.7-Plus", "风格分析 + 文章生成 同一模型"),
    ("输入单价（≤256k）", 1.6, "¥/M tokens，截图限时 8 折价"),
    ("输出单价（≤256k）", 6.4, "¥/M tokens，截图限时 8 折价"),
    ("每汉字 ≈ tokens", 1.5, "中文经验值，用于估算输出 token"),
    ("文章基础输入 tokens", 1500, "风格档案(约1000) + 写作要求(约500)"),
    ("平均单积分成本(¥)", 0.0043, "按 2000字中篇测算：输入1500+输出3000 tokens = ¥0.0216 / 5积分"),
    ("积分定位", "用量配额单位", "1 积分为用户侧配额单位，与人民币非 1:1，仅用于限制用量；真实成本按 token 另计"),
]
for i, (k, v, note) in enumerate(rows, start=3):
    ws1.cell(row=i, column=1, value=k)
    ws1.cell(row=i, column=2, value=v)
    ws1.cell(row=i, column=3, value=note)
    ws1.cell(row=i, column=1).border = thin_border
    ws1.cell(row=i, column=2).border = thin_border
    ws1.cell(row=i, column=3).border = thin_border
    ws1.cell(row=i, column=2).font = blue_font
ws1["B4"].number_format = '"¥"#,##0.00"/M"'
ws1["B5"].number_format = '"¥"#,##0.00"/M"'
ws1["B6"].number_format = "0.0"
ws1["B7"].number_format = "#,##0"
ws1["B8"].number_format = '"¥"#,##0.0000'

# ---------------- Sheet 2: 操作积分设计 ----------------
ws2 = wb.create_sheet("操作积分设计")
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 46
ws2["A1"] = "每个操作的花费积分设计"
ws2["A1"].font = title_font
ws2.merge_cells("A1:D1")

ws2["A3"] = "一、文章生成（按长度分档，长文享折扣）"
ws2["A3"].font = Font(bold=True, size=12)
ws2.merge_cells("A3:D3")
for c, h in enumerate(["长度档（字）", "消耗积分", "每千字平均积分", "说明"], 1):
    ws2.cell(row=4, column=c, value=h)
hdr(ws2, 4, 4)
article = [
    ("≤1000", 3, "基础档，单价最高"),
    ("≤2000", 5, "比 1000 字档单位更便宜（折扣）"),
    ("≤3000", 8, "锚点：更长内容单位价更低"),
    ("≤4000", 11, "之后每千字约 +3 分"),
    ("≤5000", 14, ""),
    ("≤8000", 22, "长文创作场景"),
    ("≤10000", 29, "上限，超出需联系商务"),
]
r = 5
for length, pts, note in article:
    ws2.cell(row=r, column=1, value=length)
    ws2.cell(row=r, column=2, value=pts)
    ws2.cell(row=r, column=3, value=f"=B{r}/(VALUE(SUBSTITUTE(A{r},\"≤\",\"\"))/1000)")
    ws2.cell(row=r, column=4, value=note)
    ws2.cell(row=r, column=2).font = blue_font
    r += 1
data_border(ws2, 5, r - 1, 1, 4)
for rr in range(5, r):
    ws2.cell(row=rr, column=3).number_format = "0.00"

# Section B
sb = r + 1
ws2.cell(row=sb, column=1, value="二、其他操作（固定积分）")
ws2.cell(row=sb, column=1).font = Font(bold=True, size=12)
ws2.merge_cells(f"A{sb}:D{sb}")
for c, h in enumerate(["操作", "消耗积分", "每千字平均积分", "说明"], 1):
    ws2.cell(row=sb + 1, column=c, value=h)
hdr(ws2, sb + 1, 4)
others = [
    ("风格分析", 2, "上传作品诊断六维风格，一次性动作"),
    ("段落重写", 1, "每次重写消耗，与长度无关"),
    ("风格档案编辑", 0, "免费，不计入额度"),
    ("文章保存/下载", 0, "免费，不计入额度（仅付费档开放功能）"),
]
rr = sb + 2
for name, pts, note in others:
    ws2.cell(row=rr, column=1, value=name)
    ws2.cell(row=rr, column=2, value=pts)
    ws2.cell(row=rr, column=3, value="—")
    ws2.cell(row=rr, column=4, value=note)
    ws2.cell(row=rr, column=2).font = blue_font
    rr += 1
data_border(ws2, sb + 2, rr - 1, 1, 4)

# ---------------- Sheet 3: 用户等级套餐 ----------------
ws3 = wb.create_sheet("用户等级套餐")
cols = ["等级", "月费(¥)", "月积分", "约可生成中篇(2000字)", "风格上限", "作品上限",
        "下载docx", "段落重写", "满额成本(¥)", "毛利率"]
widths = [12, 10, 10, 16, 10, 10, 12, 12, 14, 12]
for i, w in enumerate(widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3["A1"] = "用户等级与积分套餐设计"
ws3["A1"].font = title_font
ws3.merge_cells("A1:J1")
for c, h in enumerate(cols, 1):
    ws3.cell(row=3, column=c, value=h)
hdr(ws3, 3, len(cols))

tiers = [
    ("免费版", 0, 10, 2, 3, "否", "否"),
    ("基础版", 19, 60, 10, 20, "是", "是"),
    ("专业版", 49, 200, 30, 100, "是", "是"),
    ("团队版", 199, 1000, "不限", "不限", "是", "是"),
]
r = 4
for name, price, pts, styles, works, dl, rw in tiers:
    ws3.cell(row=r, column=1, value=name)
    ws3.cell(row=r, column=2, value=price)
    ws3.cell(row=r, column=3, value=pts)
    if isinstance(styles, int):
        ws3.cell(row=r, column=4, value=f"=C{r}/5")
    else:
        ws3.cell(row=r, column=4, value=styles)
    ws3.cell(row=r, column=5, value=styles if isinstance(styles, str) else styles)
    ws3.cell(row=r, column=6, value=works)
    ws3.cell(row=r, column=7, value=dl)
    ws3.cell(row=r, column=8, value=rw)
    ws3.cell(row=r, column=9, value=f"=C{r}*模型定价假设!$B$8")
    if price == 0:
        ws3.cell(row=r, column=10, value="N/A")
    else:
        ws3.cell(row=r, column=10, value=f"=(B{r}-I{r})/B{r}")
    r += 1
data_border(ws3, 4, r - 1, 1, len(cols))
for rr in range(4, r):
    ws3.cell(row=rr, column=2).number_format = "¥#,##0"
    ws3.cell(row=rr, column=3).number_format = "#,##0"
    ws3.cell(row=rr, column=4).number_format = "0"
    ws3.cell(row=rr, column=5).number_format = "0"
    ws3.cell(row=rr, column=6).number_format = "0"
    ws3.cell(row=rr, column=9).number_format = "¥#,##0.00"
    ws3.cell(row=rr, column=10).number_format = "0.0%"

note_r = r + 1
ws3.cell(row=note_r, column=1, value="注：免费版仅可生成 ≤2000字 内容；付费档开放下载/重写；团队版含成员管理与共享风格。满额成本按平均单积分成本(¥0.0043)测算。")
ws3.cell(row=note_r, column=1).font = note_font
ws3.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=10)

# ---------------- Sheet 4: 盈亏平衡测算 ----------------
ws4 = wb.create_sheet("盈亏平衡测算")
ws4["A1"] = "以「专业版 ¥49/月，200 积分」为例：不同使用率下的真实利润"
ws4["A1"].font = title_font
ws4.merge_cells("A1:F1")
for c, h in enumerate(["使用率", "实际消耗积分", "实际成本(¥)", "月收入(¥)", "毛利(¥)", "毛利率"], 1):
    ws4.cell(row=3, column=c, value=h)
hdr(ws4, 3, 6)
for i, rate in enumerate([0.2, 0.5, 1.0], start=4):
    ws4.cell(row=i, column=1, value=f"{int(rate*100)}%")
    ws4.cell(row=i, column=2, value=f"=200*{rate}")
    ws4.cell(row=i, column=3, value=f"=B{i}*模型定价假设!$B$8")
    ws4.cell(row=i, column=4, value=49)
    ws4.cell(row=i, column=5, value=f"=D{i}-C{i}")
    ws4.cell(row=i, column=6, value=f"=E{i}/D{i}")
for i in range(4, 7):
    data_border(ws4, i, i, 1, 6)
    ws4.cell(row=i, column=2).number_format = "#,##0"
    ws4.cell(row=i, column=3).number_format = "¥#,##0.00"
    ws4.cell(row=i, column=4).number_format = "¥#,##0"
    ws4.cell(row=i, column=5).number_format = "¥#,##0.00"
    ws4.cell(row=i, column=6).number_format = "0.0%"
for col in range(1, 7):
    ws4.column_dimensions[get_column_letter(col)].width = 16

# ---------------- Sheet 5: 设计说明 ----------------
ws5 = wb.create_sheet("设计说明")
ws5.column_dimensions["A"].width = 110
ws5["A1"] = "用量与额度 · 设计说明"
ws5["A1"].font = title_font
notes = [
    "一、积分定位",
    "· 积分为「用户侧用量配额单位」，不直接等于人民币，目的是与具体模型解耦，便于换模型/做促销。",
    "· 真实成本按每次调用的 input/output token 实时计算（见模型定价假设），用于我们内部核算，不展示给用户。",
    "· 用户看到的只有「本次消耗 X 积分」和「本月剩余 Y 积分」。",
    "",
    "二、文章计费折扣逻辑",
    "· 按输出字数分档：1000字=3分、2000字=5分、3000字=8分（用户给定锚点）。",
    "· 不是严格线性：前 1000 字单价最高(3.0 分/千字)，2000 字档最低(2.5 分/千字)，更长趋于 ~2.7 分/千字，相当于「写得越长越划算」，引导深度创作。",
    "· 超过 10000 字需联系商务或单独计费。",
    "",
    "三、其他操作",
    "· 风格分析 2 分（一次性诊断）；段落重写 1 分/次；风格编辑、保存、下载本身免费，仅按等级开放功能权限。",
    "",
    "四、用户等级与权限",
    "· 免费版：10 分/月，限 ≤2000字 生成，不支持下载与重写，风格≤2、作品≤3。用于体验与获客。",
    "· 基础版 ¥19：60 分，开放下载与重写，风格≤10、作品≤20。",
    "· 专业版 ¥49：200 分，全功能 + 优先队列，风格≤30、作品≤100。",
    "· 团队版 ¥199：1000 分，含成员管理、共享风格、不限风格/作品。",
    "",
    "五、盈亏测算结论",
    "· 即便用户 100% 用满额度，各档毛利率仍约 98%（大模型成本极低）。",
    "· 实际平均使用率通常 20%~50%，真实毛利更高，商业化空间充足。",
    "",
    "六、实施要点（后端）",
    "· User 表：tier、monthly_quota、used_quota、quota_resets_at。",
    "· UsageRecord 表：operation_type、input_tokens、output_tokens、points_consumed、cost_cny、model、created_at。",
    "· 生成/分析前校验剩余额度（points ≤ remaining），不足返回 402/403 友好提示。",
    "· 每月 quota_resets_at 到期自动重置 used_quota=0。",
]
for i, t in enumerate(notes, start=3):
    ws5.cell(row=i, column=1, value=t)
    if t and not t.startswith("·") and t[0].isdigit() and "." in t[:3]:
        ws5.cell(row=i, column=1).font = Font(bold=True, size=11, color="534AB7")

wb.save("outputs/墨小小_用量额度与成本测算.xlsx")
